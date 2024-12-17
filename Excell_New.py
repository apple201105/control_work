import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('список.xlsx')
ws = wb.active
row = ws.max_row
column = ws.max_column
Average_Column = 1
for j in range(1, row+1):
    sum_row = 0
    sum_int = 1
    for i in range(2, column+2):
        cell_obj = ws.cell(row=j, column=i)
        if j == 1 and (cell_obj.value is None or cell_obj.value == "Average"):
            cell_obj.value = "Average"
            Average_Column = i
            break
        elif j > 1 and cell_obj.value is not None:
            if type(cell_obj.value) is int:
                sum_row = sum_row + int(cell_obj.value)
                sum_int = sum_int + 1
        elif i == Average_Column:
            cell_obj.value = sum_row/sum_int
wb.save('список.xlsx')
wb.close()
print("Строк-",row,"Стобцов-",Average_Column-1)
df_excel = pd.read_excel('список.xlsx')
print(df_excel)
