import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('список.xlsx')
ws = wb.active
Iv =  (ws['B2'].value+ws['C2'].value+ws['D2'].value+ws['E2'].value+ws['F2'].value)/5
Pet = (ws['B3'].value+ws['C3'].value+ws['D3'].value+ws['E3'].value+ws['F3'].value)/5
Ar =  (ws['B4'].value+ws['C4'].value+ws['D4'].value+ws['E4'].value+ws['F4'].value)/5
Rod = (ws['B5'].value+ws['C5'].value+ws['D5'].value+ws['E5'].value+ws['F5'].value)/5
And = (ws['B6'].value+ws['C6'].value+ws['D6'].value+ws['E6'].value+ws['F6'].value)/5
Vas = (ws['B7'].value+ws['C7'].value+ws['D7'].value+ws['E7'].value+ws['F7'].value)/5
Jp =  (ws['B8'].value+ws['C8'].value+ws['D8'].value+ws['E8'].value+ws['F8'].value)/5
Kir = (ws['B9'].value+ws['C9'].value+ws['D9'].value+ws['E9'].value+ws['F9'].value)/5
ws['G1'] = "AM"
ws['G2'] = Iv
ws['G3'] = Pet
ws['G4'] = Ar
ws['G5'] = Rod
ws['G6'] = And
ws['G7'] = Vas
ws['G8'] = Jp
ws['G9'] = Kir
ws['H1'] = "Average"
ws['H2'] = '= AVERAGE(B2:F2)'
ws['H3'] = '= AVERAGE(B3:F3)'
ws['H4'] = '= AVERAGE(B4:F4)'
ws['H5'] = '= AVERAGE(B5:F5)'
ws['H6'] = '= AVERAGE(B6:F6)'
ws['H7'] = '= AVERAGE(B7:F7)'
ws['H8'] = '= AVERAGE(B8:F8)'
ws['H9'] = '= AVERAGE(B9:F9)'
wb.save('список.xlsx')
df_excel = pd.read_excel('список.xlsx')
print(df_excel)

