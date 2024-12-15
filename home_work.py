import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('список.xlsx')
ws = wb.active
Iv = (5+5+2+3+4)/5
Pet = (5+5+2+3+4)/5
Ar = (5+5+2+3+4)/5
Rod = (5+5+5+5+5)/5
And = (5+5+2+3+4)/5
Vas = (5+5+2+3+4)/5
jp = (5+5+2+3+4)/5
Kir = (5+5+2+3+4)/5
ws['G1'] = "arithmetic mean"
ws['G2'] = Iv
ws['G3'] = Pet
ws['G4'] = Ar
ws['G5'] = Rod
ws['G6'] = And
ws['G7'] = Vas
ws['G8'] = jp
ws['G9'] = Kir


wb.save('список.xlsx')
df_excel = pd.read_excel('список.xlsx')
print(df_excel)
