import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('список.xlsx')
ws = wb.active
row_tab = ws.max_row
column_tab = ws.max_column
Average_Column = column_tab
# по первой строке ищем первый пустой столбец, в него будем писать среднии оценки
for i in range(1, column_tab):
    cell_obj = ws.cell(row=1, column=i)
    if cell_obj.value is None or cell_obj.value == "Average":
        cell_obj.value = "Average"# название для стобца со средними оценками
        Average_Column = i
        break
for j in range(2, row_tab+1):# внешний цикл для строк
    sum_row = 0 # при смене строк, обнуляем сумму по стобцам в строке и количество слогаемых
    sum_int = 0
    for i in range(1, Average_Column+1):# внутриний цикл для стобцов
        cell_obj = ws.cell(row=j, column=i) # получаем ссылку на ячейку таблицы
        if i < Average_Column and cell_obj.value is not None: #условие для стобца его номер должен быть меньше номера для столбца среднего значения
            if type(cell_obj.value) is int:
                sum_row = sum_row + int(cell_obj.value)
                sum_int = sum_int + 1
        if i == Average_Column: # если дошли до стобца для среднего значения, то делаем расчёт в эту ячейку
            FIO = ws.cell(row=j, column=1)
            print(FIO.value, "Сумма оценок = ", sum_row, 'Количество оценок - ', sum_int, 'Средний бал = ', sum_row/sum_int)
            cell_obj.value = sum_row/sum_int
wb.save('список.xlsx')
wb.close()
print("Строк-",row_tab,"Стобцов-",Average_Column-1)
df_excel = pd.read_excel('список.xlsx')
print(df_excel)
