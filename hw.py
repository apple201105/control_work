import pandas as pd
from openpyxl import load_workbook
wb = load_workbook('список.xlsx')
ws = wb.active
row = ws.max_row
ws['G1'] = "AM"
for i in range(2, row+1):
    AM = "G"
    AM  = AM  + str(i)
    sum_row = 0
    col_list = ["B", "C", "D", "E", "F"]
    for j in range(len(col_list)):
        col_list[j] = col_list[j] + str(i)
        sum_row = sum_row + ws[col_list[j]].value
    ws[AM] = sum_row/len(col_list)
wb.save('список.xlsx')
wb.close()
df_excel = pd.read_excel('список.xlsx')
print(df_excel)
